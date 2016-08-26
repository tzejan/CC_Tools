import sys
import csv
import json
import re

sys.path.insert(0, './lib')
import openpyxl



def processLabelValues(value):
    if type(value) is not str and type(value) is not unicode:        
        return value

    stripChars = " :.*"
    returnVal = value.strip(stripChars).upper().replace(" ","")

    # special case for List of failures for the xlsx file
    if returnVal.startswith("A)"):
        returnVal = returnVal[2:]

    return returnVal

def loadLabels():
    module_labels = {}
    GSM_labels = {}
    try:
        with open("labels.json", "r") as f:
            data = json.loads(f.read())
    except IOError:
        sys.exit("FAILED!!! labels.json file not found!")

    metadata = data['metadata']
    
    # putting the label as the key, for faster query
    for datapoint in data['module_labels']:
        label = processLabelValues(datapoint["label"])
        module_labels[label] = datapoint

    # putting the label as the key, for faster query
    for datapoint in data['GSM_labels']:
        label = processLabelValues(datapoint["label"])
        GSM_labels[label] = datapoint     
         
    return module_labels, GSM_labels, metadata

def generateMPSFilename(labels):
    #this is a bit hardcoded
    #copy the header name and put here
    filename_labels = [ "Acad Year/Semester", "Module Code", "Module Name" ]
    filename_suffix = "MPS.xlsx"

    filename_values = [labels[processLabelValues(label)]['data'] for label in filename_labels]
    filename_values.append(filename_suffix)

    #get diploma name
    diploma = labels[processLabelValues("Name of Course")]['data']
    dip = re.sub(r"[\&\(\)]|\bDIPLOMA\b|\bIN\b", "", diploma).split()
    dip_short = "".join([w[0] for w in dip])
    #print diploma, dip, dip_short

    filename_values.insert(0, dip_short)

    #join up all the values, and sanitize the filename
    filename = "_".join(filename_values)
    #print filename
    keepcharacters = (' ','.','_')
    filename = "".join(c for c in filename if c.isalnum() or c in keepcharacters).rstrip()
    #print filename

    return filename

def getModuleType(csv_data):
    module_type = ""
    module_code = ""

    for row_idx, row in enumerate(csv_data):
        #print "row %d\n" % row_idx
        for col_idx, cell_value in enumerate(row):
            if cell_value == "Module Code:":
                module_code = row[col_idx+1]
                break
        if module_code:
            break;

    #what kind of module is this?
    if module_code.startswith("DMS"):
        module_type = "GSM"
    elif module_code.startswith("DM"):
        module_type = "core"

    return module_type


def generateMPS(csvFile = None):    
    module_labels, GSM_labels, metadata = loadLabels()    
    csv_data = importCSV(csvFile, module_labels)
    module_type = getModuleType(csv_data)

    print "%s module detected!" %  (module_type)
    
    labels = {}
    MPSSource = ""
    if module_type == "core":
        labels = module_labels
        MPSSource = metadata['module_MPS']
    elif module_type == "GSM":
        labels = GSM_labels
        MPSSource = metadata['GSM_MPS']
    
    if not labels:
        print "Unable to determine module type from MPCS CSV file!"
        sys.exit(1)

    grabData(csv_data, labels)
    calculateCutOffForDistinction(labels)

    dstMPSFilename = generateMPSFilename(labels)    
    populateMPS(MPSSource, labels, dstMPSFilename)
    #cleanupMPS(dstMPSFilename) # clean up the ** Module is combined with blah blah
    print "\n"
    print "SUCCESS!!! Please check the generated MPS for any errors before printing."
    print "%s was generated." % dstMPSFilename

def grabData(csv_data, labels):
    for row_idx, row in enumerate(csv_data):
        #print "row %d\n" % row_idx
        for col_idx, cell_value in enumerate(row):
            cell = processLabelValues(cell_value)
            #print "col %d %s "% (col_idx, cell)
            if cell in labels:
                if "row" in labels[cell]["csv_offset"]:
                    #print "Single Data", cell
                    labels[cell]["data"] = getSingleData(csv_data, labels, cell, row_idx, col_idx)
                elif "variable" in labels[cell]:
                    #print "Variable data", cell
                    labels[cell]["data"] = getVariableData(csv_data, labels, cell, row_idx, col_idx)
                elif "row_range_start" in labels[cell]["csv_offset"]:
                    #print "Ranged data", cell
                    labels[cell]["data"] = getRangedData(csv_data, labels, cell, row_idx, col_idx)
                if "assessment_components" in labels[cell]:
                    #print "Need assessment components" 
                    labels[cell]["assessment_components"]["data"] = getAssessmentComponents(csv_data, labels, cell, row_idx, col_idx)
                
                print "FOUND %s %s" % (cell_value, labels[cell].get("data", ""))
        #print "\n"
    #for key, val in labels.iteritems():
    #    print key, labels[key]["data"]

# the main labels
def getSingleData(csv_data, labels, label_found, location_row, location_col):
    row_offset = labels[label_found]["csv_offset"]["row"]
    col_offset = labels[label_found]["csv_offset"]["col"]
    #print csv_data[location_row+row_offset][location_col+col_offset]
    return csv_data[location_row+row_offset][location_col+col_offset]

# only the main section
def getRangedData(csv_data, labels, label_found, location_row, location_col):
    row_range_start_offset = labels[label_found]["csv_offset"]["row_range_start"]
    col_range_start_offset = labels[label_found]["csv_offset"]["col_range_start"]
    row_range_end_offset = labels[label_found]["csv_offset"]["row_range_end"]
    col_range_end_offset = labels[label_found]["csv_offset"]["col_range_end"]

    data = []
    for ro in range(row_range_start_offset, row_range_end_offset):
        row_data = csv_data[location_row+ro][location_col+col_range_start_offset:location_col+col_range_end_offset]        
        data.append(row_data)
    #print data
    return data

#these are the special cases, we will need to do something customized for these
def getVariableData(csv_data, labels, label_found, location_row, location_col):
    row_range_start_offset = labels[label_found]["csv_offset"]["row_range_start"]
    col_range_start_offset = labels[label_found]["csv_offset"]["col_range_start"]
    

    admin_num_regex = re.compile("[0-9]{6}[A-Z]")

    data = []    
    row_count = 0
    num_col = len(csv_data[location_row+row_range_start_offset]) - col_range_start_offset - location_col

    #we hack this, assume the first col contains admin number
    while len(csv_data[location_row+row_range_start_offset+row_count]) > location_col+col_range_start_offset \
        and admin_num_regex.match(csv_data[location_row+row_range_start_offset+row_count][location_col+col_range_start_offset]):

        row_data = csv_data[location_row+row_range_start_offset+row_count][location_col+col_range_start_offset:location_col+col_range_start_offset+num_col]
        #print row_data
        row_count += 1
        data.append(row_data)

    #print "num_col = %d" % num_col
    #print "len = %d" % len(csv_data[location_row+row_range_start_offset])    
    return data

#assessment components have an offset, and we assume that the assessment components headers are all upppercase
def getAssessmentComponents(csv_data, labels, label_found, location_row, location_col):
    row_range_start_offset = labels[label_found]["assessment_components"]['csv_offset']["row_range_start"]
    col_range_start_offset = labels[label_found]["assessment_components"]['csv_offset']["col_range_start"]
    asm_components_row_count = labels[label_found]["assessment_components"]['row_count']

    data = []
    col_count = 0

    #get the number of assessment components by using the headers
    headers_regex = re.compile("^[A-Z0-9]+$")

    while len(csv_data[location_row+row_range_start_offset]) > location_col+col_range_start_offset+col_count \
        and headers_regex.match(csv_data[location_row+row_range_start_offset][location_col+col_range_start_offset+col_count]):
        
        #print csv_data[location_row+row_range_start_offset][location_col+col_range_start_offset+col_count]
        col_count += 1
    
    num_asm_components = col_count
    #print num_asm_components

    # get the assessment components
    for co in range(num_asm_components):
        component = []
        for ro in range(asm_components_row_count):
            component.append(csv_data[location_row+row_range_start_offset+ro][location_col+col_range_start_offset+co])
        #print component
        data.append(component)
    
    #print data
    return data

def importCSV(csvFile, labels):
    f = open(csvFile)

    reader = csv.reader(f)

    data = []

    for row_index, row in enumerate(reader):
        row_data = []
        for column_index, cell in enumerate(row):
            #print "col ",column_index, cell,
            value = cell.strip()
            #try to parse as int or float
            #try float first because it is stricter            
            try:
                value = float(value)
            except ValueError:
                pass

            try:
                value = int(value)
            except ValueError:
                pass
                        
            #print value
            row_data.append(value)
        #print ""

        data.append(row_data)
        #print row_indexprint row_index, row

    #print data
    return data

# we take the reference MPS and then create another MPS with the populated data
def populateMPS(MPSSource, labels, dstMPSFilename):
    wb = openpyxl.load_workbook(MPSSource)
    #assume first sheet is the one we are using
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    #print sheet.max_row
    #print sheet.max_column

    #for key, val in labels.iteritems():
    #    print key
    #print "------------------"
    writeDistinctionFootnote(labels, sheet)

    for r in range(1, sheet.max_row+1):
        for c in range(1, sheet.max_column+1):
            # labels are usually bold
            if not sheet.cell(row=r, column=c).font.b:
                continue

            value = sheet.cell(row=r, column=c).value

            if not value:
                continue

            # encode everything to UTF-8    
            if value and type(value) is unicode:
                value = processLabelValues(value.encode("UTF-8"))
            
            if value in labels:
                if "row" in labels[value]["csv_offset"]:
                    #print "Single Data", cell
                    writeSingleData(sheet, labels[value].get("data", ""), labels, value, r, c)
                elif "variable" in labels[value]:
                    #print "Variable data", cell
                    writeVariableData(sheet, labels[value]["data"], labels, value, r, c)
                elif "row_range_start" in labels[value]["csv_offset"]:
                    #print "Ranged data", cell
                    writeRangedData(sheet, labels[value]["data"], labels, value, r, c)
                if "assessment_components" in labels[value]:
                    #print "Need assessment components" 
                    writeAssessmentComponents(sheet, labels[value]["assessment_components"]["data"], labels, value, r, c)
                
        
    try:
        wb.save(dstMPSFilename)
    except:
        print "\n"
        print "FAILED!!!"
        print "%s is open, please close it and run this program again." % dstMPSFilename
        sys.exit(1)

def writeSingleData(sheet, data, labels, label, location_row, location_col):
    #print "writing single data %s, %s" %(label, data)

    row_offset = labels[label]["xls_offset"]["row"]
    col_offset = labels[label]["xls_offset"]["col"]
        
    data_row = location_row+row_offset
    data_col = location_col+col_offset

    sheet.cell(row=data_row, column=data_col).value = data

def clearTheField(sheet, labels, label, location_row, location_col):
    row_range_start_offset = labels[label]["xls_offset"]["row_range_start"]
    col_range_start_offset = labels[label]["xls_offset"]["col_range_start"]

    num_physical_rows = labels[label]["current_xlsx_rows"]
    num_physical_cols = labels[label]["physical_columns"]

    for ro in range(num_physical_rows):
        for co in range(-1, num_physical_cols):

            data_row = location_row+row_range_start_offset+ro
            data_col = location_col+col_range_start_offset+co

            cell = sheet.cell(row=data_row, column=data_col)
            cell.value = None
     

def writeVariableData(sheet, data, labels, label, location_row, location_col):
    #print "writing variable data %s, %s" %(label, data)
    if not data:
        return 

    from openpyxl.styles import Side

    thinSide = Side(border_style='thin', color="FF000000")
    mediumSide = Side(border_style='medium', color="FF000000")

    row_range_start_offset = labels[label]["xls_offset"]["row_range_start"]
    col_range_start_offset = labels[label]["xls_offset"]["col_range_start"]

    #due to a bug in the template, we need to insert a column in the xlsx file
    insert_col = labels[label]["empty_data_column"]
   
    num_data_cols = len(data[0])
    num_physical_cols = labels[label]["physical_columns"]

    # the excel file is badly formatted, so we get the styles from the first line
    # we start from -1 because of numbering
    styles = []
    for co in range(-1, num_physical_cols):
        data_row = location_row+row_range_start_offset
        data_col = location_col+col_range_start_offset+co

        #get the style
        style = {}
        style['font'] = sheet.cell(row=data_row, column=data_col).font
        style['alignment'] = sheet.cell(row=data_row, column=data_col).alignment
        style['number_format'] = sheet.cell(row=data_row, column=data_col).number_format
        style['border'] = sheet.cell(row=data_row, column=data_col).border.copy(bottom=thinSide, top=thinSide)
        style['top-border'] = sheet.cell(row=data_row, column=data_col).border.copy(bottom=thinSide, top=mediumSide)
        style['bottom-border'] = sheet.cell(row=data_row, column=data_col).border.copy(bottom=mediumSide, top=thinSide)
        styles.append(style)
        
    # clear out existing data in the template MPS first
    clearTheField(sheet, labels, label, location_row, location_col)


    # need to offset the column by 1 to the left for numbering
    for ro in range(len(data)):
        for co in range(-1, num_physical_cols):

            data_row = location_row+row_range_start_offset+ro
            data_col = location_col+col_range_start_offset+co
            data_idx = co
            value = None

            if co > insert_col:
                data_idx -= 1

            #first column for numbering
            if co == -1:
                value = ro+1
            elif co != insert_col and data_idx < num_data_cols:
                value = data[ro][data_idx]
            
            
            #cos array start from 0, not -1 in co
            style_idx = co + 1

            cell = sheet.cell(row=data_row, column=data_col)
            cell.font = styles[style_idx]['font']
            cell.alignment = styles[style_idx]['alignment']
            cell.number_format = styles[style_idx]['number_format']
            if ro == 0:
                cell.border = styles[style_idx]['top-border']
            elif ro == len(data) - 1 and ro > labels[label]["current_xlsx_rows"] - 1:
                cell.border = styles[style_idx]['bottom-border']
            else:
                cell.border = styles[style_idx]['border']            
            cell.value = value

            #print cell.border.top, cell.border.bottom
    


def writeRangedData(sheet, data, labels, label, location_row, location_col):
    #print "writing ranged data %s, %s" %(label, data)

    row_range_start_offset = labels[label]["xls_offset"]["row_range_start"]
    col_range_start_offset = labels[label]["xls_offset"]["col_range_start"]
    row_range_end_offset = labels[label]["xls_offset"]["row_range_end"]
    col_range_end_offset = labels[label]["xls_offset"]["col_range_end"]
    
    if row_range_end_offset-row_range_start_offset != len(data):
        print "Number of rows differ! Check your json file"
        return

    if col_range_end_offset-col_range_start_offset != len(data[0]):
        print "Number of columns differ! Check your json file"
        return        

    for ro in range(row_range_end_offset - row_range_start_offset):
        row_data = data[ro]
        for co in range(col_range_end_offset - col_range_start_offset):
            #print row_data[co]
            data_row = location_row + row_range_start_offset + ro
            data_col = location_col + col_range_start_offset + co
            cell = sheet.cell(row=data_row, column=data_col)
            cell.value = data[ro][co]
        
    #print data
    #return data

def writeAssessmentComponents(sheet, data, labels, label, location_row, location_col):
    #print "writing assessment component data %s, %s" %(label, data)

    row_range_start_offset = labels[label]["assessment_components"]['xls_offset']["row_range_start"]
    col_range_start_offset = labels[label]["assessment_components"]['xls_offset']["col_range_start"]
    asm_components_row_count = labels[label]["assessment_components"]['row_count']
    
    max_asm_comp_count = 4
    center = openpyxl.styles.alignment.Alignment(horizontal="center")

    for r in range(asm_components_row_count):
        for c in range(max_asm_comp_count):
            #for module assessment components less than 4, we set to None
            value = None
            if c < len(data):                
                value = data[c][r]

            data_row = location_row+row_range_start_offset+r
            data_col = location_col+col_range_start_offset+c

            cell = sheet.cell(row=data_row, column=data_col)
            cell.alignment = center
            cell.value = value

def calculateCutOffForDistinction(labels):
    calKey = None    
    for key, val in labels.iteritems():
        if "calculate_cutoff" in val:
            calKey = key

    if not calKey:
        return

    srcLabel = processLabelValues(labels[calKey]["calculate_cutoff"]["src"])
    dstLabel = processLabelValues(labels[calKey]["calculate_cutoff"]["dst"])
    num_eligible = labels[srcLabel].get('data', 0)
    score_idx = labels[calKey]["calculate_cutoff"]["score_idx"]

    student_cutoff = num_eligible-1
    num_qualified = len(labels[calKey]["data"])

    if num_qualified == 0: # no student eligible, that's bad
        return #nothing to do

    if student_cutoff >= num_qualified: # fewer than the number of students qualified
        student_cutoff = num_qualified - 1 # get the last student

    score = labels[calKey]["data"][student_cutoff][score_idx]
    labels[dstLabel]["data"] = score

    #append remarks to student Recommended / Not Recommended
    remarks_idx = labels[calKey]["calculate_cutoff"]["remarks_idx"]
    shortfall = remarks_idx - len(labels[calKey]["data"][0])
    shortfall_lst = [None] * shortfall
    shortfall_lst[-1] = "Recommended"
    for row_index, row in enumerate(labels[calKey]["data"]):
        if row_index > student_cutoff:
            shortfall_lst[-1] = "Not recommended"
        row.extend(shortfall_lst)

def cleanupMPS(MPSFilename):
    wb = openpyxl.load_workbook(MPSFilename)
    #assume first sheet is the one we are using
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    remove_text = "**"   
    found = False

    for r in range(1, sheet.max_row+1):
        for c in range(1, sheet.max_column+1):
            value = sheet.cell(row=r, column=c).value
            if value and type(value) is unicode and value.startswith(remove_text):
                sheet.cell(row=r, column=c).value = None
                found = True
                break
        if found:
            break
            
    try:
        wb.save(MPSFilename)
    except:
        print "%s is open, please close it and run this program again" % MPSFilename

def writeDistinctionFootnote(labels, sheet):
    # find out which label is the one we are using
    distinction_label = None
    for key, val in labels.iteritems():
        if "footnote" in labels[key]:
            distinction_label = key

    if not distinction_label:
        return
    #see if we need additional rows, if we need, then we need to move the footnote
    #print len(labels[distinction_label]['data'])
    #print labels[distinction_label]['current_xlsx_rows']
    row_offset = len(labels[distinction_label]['data']) - labels[distinction_label]['current_xlsx_rows']

    if row_offset < 1:
        return


    footnote_label = processLabelValues(labels[distinction_label]['footnote'].encode("UTF-8"))
    #print footnote_label

    # we need to add rows, find out where is the existing cells
    for r in range(1, sheet.max_row+1):
        for c in range(1, sheet.max_column+1):
            value = sheet.cell(row=r, column=c).value
            if value and type(value) is unicode and processLabelValues(value.encode("UTF-8")) == footnote_label:
                #print sheet.cell(row=r, column=c).value, r, c
                # we make copies of these data down
                for i in range(labels[footnote_label]['csv_offset']['row_range_end'] - labels[footnote_label]['csv_offset']['row_range_start']):
                    sheet.cell(row=r+i+row_offset, column=c).value = sheet.cell(row=r+i, column=c).value
                    sheet.cell(row=r+i+row_offset, column=c).font = sheet.cell(row=r+i, column=c).font
                return
       
          
def test():
	
	wb = Workbook()

	# grab the active worksheet
	ws = wb.active

	# Data can be assigned directly to cells
	ws['A1'] = 42

	# Rows can also be appended
	ws.append([1, 2, 3])

	# Python types will automatically be converted
	import datetime
	ws['A2'] = datetime.datetime.now()

	# Save the file
	wb.save("sample.xlsx")

if __name__ == "__main__":
    MPCSFile = "MPCS_DMDF12.CSV"
    MPCSFile = r"C:\Users\simtj\OneDrive\Work\Course\DMS342 Effective Communications for Better Relationships\DMS342 MPCS_DMDF06.CSV"
    if len(sys.argv) > 1:        
        MPCSFile = sys.argv[1]
    else:
        raw_input( "Enter the filename of your Module Performance Comparison Summary file or drag the file and drop it on top of this executable file to begin." )
        sys.exit(0)

    generateMPS(MPCSFile)
    